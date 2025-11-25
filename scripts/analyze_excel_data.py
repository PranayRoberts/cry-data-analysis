import openpyxl
import pandas as pd
import json
import os
from pathlib import Path
import math

# Define paths - use absolute paths
script_dir = Path(__file__).parent.parent
excel_dir = script_dir / 'excel-data'
output_dir = script_dir / 'public' / 'data'
output_dir.mkdir(parents=True, exist_ok=True)

# Map of dataset pairs
dataset_pairs = {
    'anganwadi': [
        'anganwadi-centre-information_2024.xlsx',
        'anganwadi-centre-information_Report_2023.xlsx'
    ],
    'child_annual': [
        'Child_Annual_2023.xlsx',
        'child-annual-information_Report_2024.xlsx'
    ],
    'child_education': [
        'Child_education_Consolidated-2023.xlsx',
        'child-education_Consolidated_2024.xlsx'
    ],
    'school': [
        'school-level-information_2023.xlsx',
        'school-level-information_2024.xlsx'
    ]
}

def get_sheet_info(file_path):
    """Extract sheet names and row/col counts"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        info = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = ws.max_row
            max_col = ws.max_column
            # Get first few rows for schema
            data = []
            for row in ws.iter_rows(min_row=1, max_row=min(5, max_row), values_only=True):
                data.append(row)
            info[sheet_name] = {
                'rows': max_row,
                'cols': max_col,
                'sample': data
            }
        return info
    except Exception as e:
        return {'error': str(e)}

def analyze_all_datasets():
    """Analyze all dataset pairs"""
    analysis = {}
    
    for pair_name, files in dataset_pairs.items():
        analysis[pair_name] = {}
        for file in files:
            file_path = excel_dir / file
            if file_path.exists():
                print(f"Analyzing {file}...")
                info = get_sheet_info(file_path)
                analysis[pair_name][file] = info
            else:
                analysis[pair_name][file] = {'error': 'File not found'}
    
    return analysis

def convert_excel_to_json(file_path, sheet_name=None):
    """Convert Excel sheet to JSON"""
    try:
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(file_path)
        
        # Replace NaN/inf with None for JSON serialization
        df = df.where(pd.notna(df), None)
        
        # Convert to records and handle any remaining NaN/inf
        records = df.to_dict(orient='records')
        records = clean_nan_values(records)
        return records
    except Exception as e:
        return {'error': str(e)}

def clean_nan_values(obj):
    """Recursively replace NaN and inf with None in nested structures"""
    if isinstance(obj, dict):
        return {k: clean_nan_values(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_nan_values(item) for item in obj]
    elif isinstance(obj, float):
        # Check for NaN or inf
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    return obj

def main():
    print("=" * 80)
    print("ANALYZING EXCEL DATASETS")
    print("=" * 80)
    
    analysis = analyze_all_datasets()
    
    # Print analysis
    print("\nDATASET STRUCTURE ANALYSIS:")
    print("=" * 80)
    print(json.dumps(analysis, indent=2, default=str))
    
    # Save analysis to file
    with open(output_dir / 'schema_analysis.json', 'w', encoding='utf-8') as f:
        json.dump(analysis, f, indent=2, default=str)
    
    print("\n" + "=" * 80)
    print("Schema analysis saved to public/data/schema_analysis.json")
    print("=" * 80)
    
    # Convert all sheets to JSON for inspection
    print("\nConverting sheets to JSON...")
    for pair_name, files in dataset_pairs.items():
        pair_data = {}
        for file in files:
            file_path = excel_dir / file
            if file_path.exists():
                try:
                    # Try to get all sheets
                    xls = pd.ExcelFile(file_path)
                    for sheet in xls.sheet_names:
                        print(f"  Converting {file} sheet: {sheet}")
                        json_data = convert_excel_to_json(file_path, sheet)
                        pair_data[f"{file}_{sheet}"] = json_data
                except Exception as e:
                    print(f"  Error: {e}")
        
        # Save pair data
        if pair_data:
            output_file = output_dir / f'{pair_name}_data.json'
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(pair_data, f, indent=2)
            print(f"  Saved: {output_file}")

if __name__ == '__main__':
    main()
